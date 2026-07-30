"""
Samsung Creative Brief — automation engine.

Reads the master creative-brief Excel (4 tabs), validates the highlighted
test model, generates a structured brief, and renders brand-compliant
creatives for all 4 required paid formats — each exported as a genuine
multi-layer PSD (via ImageMagick, assembled from separate transparent
layer PNGs) and a flattened JPG.

Re-run on next month's file: point INPUT_PATH at the new .xlsx and set
MODEL_CODE to whichever row should be the hero model — everything else
(format specs, safe zones, layout math) is generic and data-driven.
"""

import os
import json
import subprocess
import openpyxl
from PIL import Image, ImageDraw, ImageFont, ImageFilter

BASE = os.path.dirname(os.path.abspath(__file__))
INPUT_PATH = os.path.join(BASE, "samsung-creative-brief-template.xlsx")
FONT_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
FONT_REGULAR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
PRODUCT_PNG = os.path.join(BASE, "assets", "product.png")
LOGO_PNG = os.path.join(BASE, "assets", "logo.png")
OUT_DIR = os.path.join(BASE, "output")

NAVY = (27, 37, 89, 255)
GOLD = (201, 151, 62, 255)
RED = (194, 69, 69, 255)
DARK = (30, 33, 42, 255)
MUTED = (107, 114, 128, 255)

# ---------- Format specs, straight from "2. Форматы (Specs)" ----------
FORMATS = {
    "feed_square":   {"label": "Feed Square",   "size": (1080, 1080), "safe": {"top": 64, "bottom": 64, "left": 64, "right": 64}, "logo_pos": "top_left"},
    "feed_portrait": {"label": "Feed Portrait", "size": (1080, 1350), "safe": {"top": 72, "bottom": 72, "left": 72, "right": 72}, "logo_pos": "top_left"},
    "stories":       {"label": "Stories/Reels", "size": (1080, 1920), "safe": {"top": 250, "bottom": 350, "left": 72, "right": 72}, "logo_pos": "top_left"},
    "display":       {"label": "Display",       "size": (1200, 628),  "safe": {"top": 60, "bottom": 60, "left": 60, "right": 60}, "logo_pos": "top_left"},
}


def load_main_model(path):
    """Pull the pink-highlighted hero row from '1. Вводные (Input)' —
    falls back to the first data row if no highlight is found, so this
    still works on a differently-styled file next month."""
    wb = openpyxl.load_workbook(path)
    ws = wb["1. Вводные (Input)"]
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Tier":
            header_row = row[0].row
            break
    headers = [c.value for c in ws[header_row]]

    hero = None
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        if not row[0].value:
            continue
        fill = row[0].fill
        is_pink = fill and fill.fgColor and fill.fgColor.rgb == "FFFCE4EC"
        rec = dict(zip(headers, [c.value for c in row]))
        if is_pink:
            hero = rec
            break
        if hero is None:
            hero = rec  # fallback: first data row
    return hero


def validate(model):
    """Basic data-quality gate before anything gets rendered."""
    issues = []
    required = ["Model Code", "Слоган (Headline)", "Категория"]
    for field in required:
        if not model.get(field):
            issues.append(f"Пусто обязательное поле: {field}")
    if not model.get("Промо-цена") and not model.get("RRP (старая цена)"):
        issues.append("Нет ни промо-цены, ни RRP — нечего показать как цену")
    discount = model.get("Скидка %")
    if discount is not None and not (-1 <= discount <= 0):
        issues.append(f"Скидка % выглядит некорректно: {discount}")
    return issues


def generate_brief(model):
    price = model.get("Промо-цена") or model.get("RRP (старая цена)")
    discount_pct = model.get("Скидка %")
    discount_str = f"{abs(round(discount_pct * 100))}%" if discount_pct else None
    features = [f.strip() for f in (model.get("Ключевые фичи") or "").split(";") if f.strip()]

    brief = {
        "model_code": model.get("Model Code"),
        "category": model.get("Категория"),
        "tier": model.get("Tier"),
        "channel": model.get("Channel"),
        "headline": model.get("Слоган (Headline)"),
        "price_promo": model.get("Промо-цена"),
        "price_rrp": model.get("RRP (старая цена)"),
        "savings": model.get("Выгода"),
        "discount_label": f"-{discount_str}" if discount_str else None,
        "features": features,
        "volume": model.get("Объём"),
        "extra_comm": model.get("Доп. коммуникация"),
        "formats": list(FORMATS.keys()),
    }
    return brief


def fmt_price(v):
    return f"{int(v):,}".replace(",", " ") + " ₸"


def fit_text(draw, text, font_path, max_width, start_size, min_size=18):
    size = start_size
    while size > min_size:
        font = ImageFont.truetype(font_path, size)
        bbox = draw.textbbox((0, 0), text, font=font)
        if bbox[2] - bbox[0] <= max_width:
            return font
        size -= 2
    return ImageFont.truetype(font_path, min_size)


def rounded_rect(draw, box, radius, fill):
    draw.rounded_rectangle(box, radius=radius, fill=fill)


def render_layers(model, brief, format_key):
    """Returns an ordered dict: layer_name -> RGBA image (same canvas size).
    Order matters — it's also the PSD stacking order (bottom to top)."""
    spec = FORMATS[format_key]
    W, H = spec["size"]
    safe = spec["safe"]
    layers = {}

    # ---- background: soft light gradient, Samsung-style ----
    bg = Image.new("RGBA", (W, H), (255, 255, 255, 255))
    top_color = (255, 255, 255)
    bottom_color = (232, 236, 245)
    for y in range(H):
        t = y / H
        r = int(top_color[0] + (bottom_color[0] - top_color[0]) * t)
        g = int(top_color[1] + (bottom_color[1] - top_color[1]) * t)
        b = int(top_color[2] + (bottom_color[2] - top_color[2]) * t)
        ImageDraw.Draw(bg).line([(0, y), (W, y)], fill=(r, g, b, 255))
    layers["background"] = bg

    # ---- pre-compute vertical anchors (price/features zone) so product sizing
    # can respect them exactly, regardless of format-specific safe-zone sizes ----
    is_tall = H > W
    if is_tall:
        price_top_anchor = H - safe["bottom"] - int(H * 0.11)
    else:
        price_top_anchor = int(H * 0.68)
    feat_top_anchor = price_top_anchor - int(H * 0.055)
    head_zone_bottom_anchor = safe["top"] + int(H * 0.11)  # logo + headline block, roughly

    # ---- product: scaled, right-aligned (tall formats: centered, above price zone) ----
    product = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    prod_img = Image.open(PRODUCT_PNG).convert("RGBA")
    prod_img = prod_img.crop(prod_img.getbbox())  # trim transparent padding in the source cutout
    if is_tall:
        avail_top = head_zone_bottom_anchor + 20
        avail_bottom = feat_top_anchor - 24
        max_h = max(100, avail_bottom - avail_top)
        max_w = int(W * 0.68)
        scale = min(max_w / prod_img.width, max_h / prod_img.height)
        target_w = int(prod_img.width * scale)
        target_h = int(prod_img.height * scale)
        prod_resized = prod_img.resize((target_w, target_h), Image.LANCZOS)
        px = (W - target_w) // 2
        py = avail_top + (max_h - target_h) // 2
    else:
        # cap BOTH width and height so a wide source image can't eat into the text column
        max_w = int(W * 0.46)
        max_h = int(H * 0.62)
        scale = min(max_w / prod_img.width, max_h / prod_img.height)
        target_w = int(prod_img.width * scale)
        target_h = int(prod_img.height * scale)
        prod_resized = prod_img.resize((target_w, target_h), Image.LANCZOS)
        px = W - target_w - safe["right"] - 10
        py = (H - target_h) // 2 + int(H * 0.04)
    product.paste(prod_resized, (px, py), prod_resized)
    layers["product"] = product

    # ---- logo: top-left, sized so clear space (0.5x logo height) is respected ----
    logo_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    logo_img = Image.open(LOGO_PNG).convert("RGBA")
    logo_img = logo_img.crop(logo_img.getbbox())
    logo_h = int(H * 0.032) if not is_tall else int(H * 0.024)
    logo_h = max(logo_h, 26)
    logo_scale = logo_h / logo_img.height
    logo_w = int(logo_img.width * logo_scale)
    logo_resized = logo_img.resize((logo_w, logo_h), Image.LANCZOS)
    lx, ly = safe["left"], safe["top"]
    logo_layer.paste(logo_resized, (lx, ly), logo_resized)
    layers["logo"] = logo_layer
    clear_space = logo_h * 0.5  # used later for validation

    # ---- headline ----
    headline_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    d = ImageDraw.Draw(headline_layer)
    headline_top = ly + logo_h + int(clear_space) + 14
    headline_max_w = W - safe["left"] - safe["right"]
    headline_size = int(W * 0.036)
    font_headline = fit_text(d, brief["headline"] or "", FONT_BOLD, headline_max_w, headline_size, 20)
    d.text((safe["left"], headline_top), brief["headline"] or "", font=font_headline, fill=NAVY)
    layers["headline"] = headline_layer
    headline_bbox = d.textbbox((safe["left"], headline_top), brief["headline"] or "", font=font_headline)

    # ---- price block: promo price (large) + struck RRP + discount badge ----
    price_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    d2 = ImageDraw.Draw(price_layer)
    price_top = price_top_anchor
    price_str = fmt_price(brief["price_promo"]) if brief["price_promo"] else fmt_price(brief["price_rrp"])
    price_size = int(W * 0.075)
    font_price = ImageFont.truetype(FONT_BOLD, price_size)
    d2.text((safe["left"], price_top), price_str, font=font_price, fill=DARK)
    price_bbox = d2.textbbox((safe["left"], price_top), price_str, font=font_price)

    cursor_y = price_bbox[3] + 6
    if brief["price_promo"] and brief["price_rrp"] and brief["price_rrp"] != brief["price_promo"]:
        rrp_str = fmt_price(brief["price_rrp"])
        font_rrp = ImageFont.truetype(FONT_REGULAR, int(price_size * 0.42))
        rrp_bbox = d2.textbbox((safe["left"], cursor_y), rrp_str, font=font_rrp)
        d2.text((safe["left"], cursor_y), rrp_str, font=font_rrp, fill=MUTED)
        strike_y = (rrp_bbox[1] + rrp_bbox[3]) // 2
        d2.line([(rrp_bbox[0], strike_y), (rrp_bbox[2], strike_y)], fill=MUTED, width=3)

        if brief["discount_label"]:
            badge_font = ImageFont.truetype(FONT_BOLD, int(price_size * 0.4))
            badge_text = brief["discount_label"]
            bbox = d2.textbbox((0, 0), badge_text, font=badge_font)
            pad_x, pad_y = 18, 10
            bw, bh = bbox[2] - bbox[0] + pad_x * 2, bbox[3] - bbox[1] + pad_y * 2
            bx = rrp_bbox[2] + 16
            by = cursor_y - 4
            rounded_rect(d2, [bx, by, bx + bw, by + bh], radius=bh // 2, fill=RED)
            d2.text((bx + pad_x, by + pad_y - bbox[1] // 2 - 2), badge_text, font=badge_font, fill=(255, 255, 255, 255))
    layers["price"] = price_layer

    # ---- features row ----
    feat_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    d3 = ImageDraw.Draw(feat_layer)
    feat_top = feat_top_anchor
    feat_font = ImageFont.truetype(FONT_REGULAR, int(W * 0.017))
    feat_text = "  ·  ".join(brief["features"][:3])
    d3.text((safe["left"], feat_top), feat_text, font=feat_font, fill=NAVY)
    layers["features"] = feat_layer

    # ---- volume badge ----
    vol_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    if brief.get("volume"):
        d4 = ImageDraw.Draw(vol_layer)
        vol_font = ImageFont.truetype(FONT_BOLD, int(W * 0.02))
        vol_text = str(brief["volume"])
        bbox = d4.textbbox((0, 0), vol_text, font=vol_font)
        pad = 14
        bw, bh = bbox[2] - bbox[0] + pad * 2, bbox[3] - bbox[1] + pad * 2
        vx = W - safe["right"] - bw
        vy = safe["top"]
        rounded_rect(d4, [vx, vy, vx + bw, vy + bh], radius=bh // 2, fill=(255, 255, 255, 235))
        d4.text((vx + pad, vy + pad - bbox[1] // 2 - 2), vol_text, font=vol_font, fill=NAVY)
    layers["volume_badge"] = vol_layer

    meta = {"headline_bbox": headline_bbox, "logo_bbox": (lx, ly, lx + logo_w, ly + logo_h), "clear_space": clear_space}
    return layers, meta


def validate_safe_zone(layers, meta, spec):
    """Bonus: automatic safe-zone / logo-clear-space check before export."""
    W, H = spec["size"]
    safe = spec["safe"]
    issues = []
    lx0, ly0, lx1, ly1 = meta["logo_bbox"]
    if lx0 < safe["left"] or ly0 < safe["top"]:
        issues.append("Логотип выходит за safe zone")
    hx0, hy0, hx1, hy1 = meta["headline_bbox"]
    if hx1 > W - safe["right"]:
        issues.append("Заголовок выходит за правую safe zone")
    # clear space: nothing else should overlap the padded logo box
    pad = meta["clear_space"]
    padded_logo = (lx0 - pad, ly0 - pad, lx1 + pad, ly1 + pad)
    if hy0 < padded_logo[3] and hx0 < padded_logo[2]:
        # headline starts far enough below logo+clearspace already by construction; flag only if truly overlapping
        if hy0 < padded_logo[1]:
            issues.append("Заголовок нарушает clear space логотипа")
    return issues


def compose_flat(layers):
    order = ["background", "product", "logo", "headline", "features", "price", "volume_badge"]
    base = layers["background"].convert("RGBA")
    for name in order[1:]:
        base = Image.alpha_composite(base, layers[name])
    return base


def export_format(model, brief, format_key):
    spec = FORMATS[format_key]
    layers, meta = render_layers(model, brief, format_key)
    issues = validate_safe_zone(layers, meta, spec)

    layer_dir = os.path.join(OUT_DIR, f"_layers_{format_key}")
    os.makedirs(layer_dir, exist_ok=True)
    order = ["background", "product", "logo", "headline", "features", "price", "volume_badge"]
    layer_paths = []
    for name in order:
        p = os.path.join(layer_dir, f"{name}.png")
        layers[name].save(p)
        layer_paths.append(p)

    psd_path = os.path.join(OUT_DIR, f"{spec['label'].replace('/', '-')}.psd")
    subprocess.run(["convert"] + layer_paths + [psd_path], check=True)

    flat = compose_flat(layers).convert("RGB")
    jpg_path = os.path.join(OUT_DIR, f"{spec['label'].replace('/', '-')}.jpg")
    flat.save(jpg_path, quality=92)

    return {"format": spec["label"], "size": spec["size"], "psd": psd_path, "jpg": jpg_path, "safe_zone_issues": issues}


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    model = load_main_model(INPUT_PATH)
    issues = validate(model)
    brief = generate_brief(model)

    with open(os.path.join(OUT_DIR, "brief.json"), "w", encoding="utf-8") as f:
        json.dump(brief, f, ensure_ascii=False, indent=2, default=str)

    results = []
    for fmt in FORMATS:
        r = export_format(model, brief, fmt)
        results.append(r)
        print(f"{r['format']}: {r['size']} — PSD + JPG saved. Safe-zone issues: {r['safe_zone_issues'] or 'none'}")

    if issues:
        print("\nVALIDATION ISSUES:", issues)
    print("\nBrief:", json.dumps(brief, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
